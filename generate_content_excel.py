import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F2937")   # dark slate
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL      = PatternFill("solid", fgColor="F3F4F6")   # light gray
BORDER_SIDE   = Side(style="thin", color="D1D5DB")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE,  bottom=BORDER_SIDE)
BODY_FONT     = Font(size=10)
WRAP          = Alignment(wrap_text=True, vertical="top")
CENTER        = Alignment(horizontal="center", vertical="top")

def style_sheet(ws, col_widths):
    # Header row
    for cell in ws[1]:
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = CENTER
        cell.border  = CELL_BORDER
    ws.row_dimensions[1].height = 24

    # Body rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            cell.font      = BODY_FONT
            cell.border    = CELL_BORDER
            cell.alignment = WRAP
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 48

    # Column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"


# ── 1. Events ────────────────────────────────────────────────────────────────
ws_events = wb.active
ws_events.title = "Events"

events_headers = ["Event Name", "Date", "Year", "Category",
                  "Time", "Location", "Status", "Description", "CTA URL"]
ws_events.append(events_headers)

events_data = [
    ["Rivertowns Chamber Annual Gala 2025", "Jun 14", "2025", "Fundraiser",
     "6:00 PM – 10:00 PM", "Rivertowns Square, Dobbs Ferry, NY", "Upcoming",
     "Join us for an evening celebrating the businesses and people of the Rivertowns. "
     "The Ardsley Carwash is proud to be a sponsor of this annual celebration of local commerce and community.",
     "https://rivertownschamber.com/gala-2026/"],
    ["Ardsley Independence Day Celebration", "Jul 4", "2025", "Festival",
     "All Day", "Ardsley, NY", "Upcoming",
     "Celebrate the Fourth of July with your Ardsley neighbors. "
     "Come find us at the festivities — free wash coupons for the community while supplies last!",
     "index.html#booking"],
    ["Ardsley Little League Opening Day 2025", "Apr 12", "2025", "Sponsorship",
     "9:00 AM – 12:00 PM", "Ardsley High School, Ardsley, NY", "Past",
     "We were proud to sponsor Opening Day for Ardsley Little League. "
     "A great morning cheering on the next generation of players with families from across the community.",
     ""],
    ["Ardsley PTA Fundraiser Night", "Mar 8", "2025", "School",
     "7:00 PM – 9:00 PM", "Ardsley Middle School, Ardsley, NY", "Past",
     "We donated wash packages to the Ardsley PTA fundraiser auction, "
     "supporting local school programs and the families who make them possible.",
     ""],
    ["Thanksgiving Wash & Give Drive", "Nov 24", "2024", "Fundraiser",
     "All Day", "639 Saw Mill River Rd, Ardsley, NY", "Past",
     "For every wash purchased on Thanksgiving eve, we donated a portion of proceeds to the local food bank. "
     "Thank you to everyone who showed up — we raised over $800 for neighbors in need.",
     ""],
]
for row in events_data:
    ws_events.append(row)

style_sheet(ws_events, [36, 10, 8, 14, 20, 36, 12, 60, 40])


# ── 2. Teams We Sponsor ──────────────────────────────────────────────────────
ws_teams = wb.create_sheet("Teams We Sponsor")

teams_headers = ["Team Name", "Description", "Logo Asset", "Website / Social Link", "Link Label"]
ws_teams.append(teams_headers)

teams_data = [
    ["Ardsley Little League",
     "Supporting young ballplayers in Ardsley since the very beginning. "
     "We're proud sponsors of the league that builds teamwork, grit, and community spirit one inning at a time.",
     "assets/partners/ardsley-little-league.png",
     "https://www.ardsleylittleleague.org/",
     "Visit Website"],
    ["Irvington Bulldogs",
     "Backing the Irvington Bulldog Booster Club and the dedicated athletes who wear the blue and white. "
     "Every win on the field is a win for our community.",
     "assets/partners/irvington-bulldog.png",
     "https://irvingtonbulldogboosterclub.com/home",
     "Visit Website"],
    ["DFYLL (Dobbs Ferry Youth Lacrosse League)",
     "Supporting Dobbs Ferry Youth Lacrosse League and the passionate players and families "
     "who make the program thrive season after season.",
     "assets/partners/dfyll.png",
     "https://www.dfyll.com/",
     "Visit Website"],
    ["PLT Racing",
     "Proud sponsors of PLT Racing — keeping the competitive spirit alive on the track. "
     "Follow their season and see what local passion looks like at full throttle.",
     "assets/partners/pltracing.jpg",
     "https://www.instagram.com/pltracing/",
     "Follow on Instagram"],
]
for row in teams_data:
    ws_teams.append(row)

style_sheet(ws_teams, [36, 60, 36, 44, 18])


# ── 3. Community Partnerships ────────────────────────────────────────────────
ws_partners = wb.create_sheet("Community Partnerships")

partners_headers = ["Organization Name", "Description", "Logo Asset", "Website Link"]
ws_partners.append(partners_headers)

partners_data = [
    ["Rivertowns Chamber of Commerce",
     "Connecting the businesses and people of the Rivertowns — Dobbs Ferry, Ardsley, Hastings, "
     "Irvington, and Tarrytown. We're proud members and sponsors of their annual Gala.",
     "assets/partners/rivertowns-chamber.png",
     "https://rivertownschamber.com/gala-2026/"],
    ["Greenburgh Chamber of Commerce",
     "Supporting local commerce in Greenburgh and the broader Westchester community. "
     "Together, we help small businesses and neighborhoods thrive.",
     "assets/partners/greenburgh-chamber.png",
     "https://www.greenburghchamber.com/home"],
]
for row in partners_data:
    ws_partners.append(row)

style_sheet(ws_partners, [36, 60, 36, 44])


# ── 4. Services ──────────────────────────────────────────────────────────────
ws_services = wb.create_sheet("Services")

services_headers = ["Service Name", "Category", "Price", "Price Note",
                    "Featured / Tag", "What's Included"]
ws_services.append(services_headers)

services_data = [
    # Wash Packages
    ["Regular Wash", "Wash Package", "$23.06", "+ tax", "",
     "Interior vacuum · Exterior wash · Towel dry · Interior dash cleaned · Windows cleaned"],
    ["Deluxe Wash", "Wash Package", "$26.76", "+ tax", "Popular",
     "Everything in Regular + Undercarriage wash · Tire shine"],
    ["Ultimate Wash", "Wash Package", "$28.60", "+ tax", "",
     "Everything in Deluxe + Hot Shine Carnauba Wax · Rubber mat wash · RAIN-X surface protectant"],
    # Exterior Options
    ["Exterior Only", "Exterior Option", "$14.76", "+ tax", "",
     "Exterior wash · Towel dry included"],
    ["Ultimate Exterior", "Exterior Option", "$20.30", "+ tax", "",
     "Towel dry · Tire shine · Triple foam wash · Hot Shine Carnauba Wax · RAIN-X surface protectant"],
    # A La Carte
    ["Tire Shine & Rims", "A La Carte Add-On", "$4.00", "", "", ""],
    ["4 Plastic Mats Cleaned", "A La Carte Add-On", "$4.00", "", "", ""],
    ["Carpet Mats Shampoo", "A La Carte Add-On", "$5.00 ea", "Per mat", "", ""],
    ["Trunk Vacuum", "A La Carte Add-On", "$4.00", "", "", ""],
    ["Interior Armor All", "A La Carte Add-On", "$20.00", "", "", ""],
    ["Express Spray Wax", "A La Carte Add-On", "$20.00", "", "", ""],
    ["Vacuum Only", "A La Carte Add-On", "$15.00", "", "", ""],
    # Hand Wash
    ["Exterior Hand Wash", "Hand Wash", "$30.00", "", "", ""],
    ["Hand Wash w/ Vacuum & Tire Shine", "Hand Wash", "$50.00", "", "",
     "Prices may vary for SUVs, pick-up trucks, and vans"],
    # Memberships
    ["Regular Monthly Car Wash", "Membership", "$50.74", "/mo",
     "Unlimited regular washes",
     "Interior vacuum · Exterior wash · Towel dry · Interior dash cleaned · Windows cleaned"],
    ["Ultimate Monthly Car Wash", "Membership", "$64.58", "/mo",
     "Includes everything in Regular",
     "Everything in Regular + Undercarriage wash · Tire shine · Rubber mat wash · RAIN surface protectant applied"],
]
for row in services_data:
    ws_services.append(row)

style_sheet(ws_services, [34, 20, 12, 12, 28, 60])


# ── 5. Reviews ───────────────────────────────────────────────────────────────
ws_reviews = wb.create_sheet("Reviews")

reviews_headers = ["Reviewer Name", "Location", "Star Rating", "Review Text"]
ws_reviews.append(reviews_headers)

reviews_data = [
    ["Michael R.", "Ardsley, NY", 5,
     "Been bringing my car here every week for five years. The team always does a thorough job "
     "and the prices are more than fair. Best car wash in Westchester, hands down."],
    ["Sarah L.", "Dobbs Ferry, NY", 5,
     "I need my car looking sharp for client meetings and Ardsley never lets me down. "
     "The Ultimate Wash with RAIN-X keeps my car showroom-ready all week long."],
    ["David K.", "Irvington, NY", 5,
     "I'm particular about my car and these guys get it. Their hand wash service is meticulous "
     "— every inch is spotless. Worth every penny for the attention to detail."],
]
for row in reviews_data:
    ws_reviews.append(row)

style_sheet(ws_reviews, [20, 20, 14, 70])


# ── 6. FAQs ──────────────────────────────────────────────────────────────────
ws_faq = wb.create_sheet("FAQs")

faq_headers = ["#", "Question", "Answer"]
ws_faq.append(faq_headers)

faq_data = [
    [1, "Do I need an appointment?",
     "Nope! We're a drive-in car wash — just pull up whenever you're ready. "
     "For detailing services, we recommend booking ahead so we can reserve your spot."],
    [2, "How long does a wash take?",
     "Our regular wash takes about 15–20 minutes. The Deluxe and Ultimate packages run 20–30 minutes. "
     "We're all about speed without cutting corners."],
    [3, "What payment methods do you accept?",
     "We accept cash, all major credit/debit cards, Apple Pay, and Google Pay. "
     "Memberships are charged monthly to your card on file."],
    [4, "Can I cancel my membership anytime?",
     "Absolutely. No contracts, no cancellation fees. You can cancel your unlimited wash membership "
     "at any time — just let us know before your next billing cycle."],
    [5, "Do you wash SUVs and trucks?",
     "Yes! We handle all vehicle sizes — sedans, SUVs, pick-up trucks, and vans. "
     "Prices may vary slightly for larger vehicles on some services."],
    [6, "What are your hours?",
     "We're open Monday–Saturday, 8:00 AM to 5:00 PM, and Sunday 8:00 AM to 4:00 PM. "
     "Weather permitting — we'll post any closures on our social media."],
]
for row in faq_data:
    ws_faq.append(row)

style_sheet(ws_faq, [6, 40, 70])


# ── Save ──────────────────────────────────────────────────────────────────────
output_path = r"c:\Users\JL\Documents\ardlseycarwash\ardsley_carwash_content.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
